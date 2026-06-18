from openpyxl import load_workbook

print("Opening workbook...")

wb = load_workbook(
    "data/sample_data_exceptions_type.xlsx",
    read_only=True
)

print("Workbook opened")
print(wb.sheetnames)